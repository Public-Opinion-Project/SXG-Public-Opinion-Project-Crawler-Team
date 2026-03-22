# -*- coding: utf-8 -*-
# Copyright (c) 2025 relakkes@gmail.com
#
# This file is part of MediaCrawler project.
# Repository: https://github.com/NanmiCoder/MediaCrawler/blob/main/store/excel_store_base.py
# GitHub: https://github.com/NanmiCoder
# Licensed under NON-COMMERCIAL LEARNING LICENSE 1.1
#
# 声明：本代码仅供学习和研究目的使用。使用者应遵守以下原则：
# 1. 不得用于任何商业用途。
# 2. 使用时应遵守目标平台的使用条款和robots.txt规则。
# 3. 不得进行大规模爬取或对平台造成运营干扰。
# 4. 应合理控制请求频率，避免给目标平台带来不必要的负担。
# 5. 不得用于任何非法或不当的用途。
#
# 详细许可条款请参阅项目根目录下的LICENSE文件。
# 使用本代码即表示您同意遵守上述原则和LICENSE中的所有条款。

# 声明:本代码仅供学习和研究目的使用。使用者应遵守以下原则:
# 1. 不得用于任何商业用途。
# 2. 使用时应遵守目标平台的使用条款和robots.txt规则。
# 3. 不得进行大规模爬取或对平台造成运营干扰。
# 4. 应合理控制请求频率,避免给目标平台带来不必要的负担。
# 5. 不得用于任何非法或不当的用途。
#
# 详细许可条款请参阅项目根目录下的LICENSE文件。
# 使用本代码即表示您同意遵守上述原则和LICENSE中的所有条款。

"""
Excel存储基础实现
提供爬取数据的Excel导出功能,带有格式化的表格
"""

import threading
from datetime import datetime
from typing import Dict, List, Any
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from base.base_crawler import AbstractStore
from tools import utils
import config


class ExcelStoreBase(AbstractStore):
    """
    Excel存储实现基类
    提供带有多工作表(内容、评论、创作者)的格式化Excel导出
    使用单例模式在多次存储调用中维护状态
    """

    # 类级别单例管理
    _instances: Dict[str, "ExcelStoreBase"] = {}
    _lock = threading.Lock()

    @classmethod
    def get_instance(cls, platform: str, crawler_type: str) -> "ExcelStoreBase":
        """
        获取或创建给定平台和爬虫类型的单例实例

        Args:
            platform: 平台名称 (xhs, dy, ks等)
            crawler_type: 爬虫类型 (search, detail, creator)

        Returns:
            ExcelStoreBase实例
        """
        key = f"{platform}_{crawler_type}"
        with cls._lock:
            if key not in cls._instances:
                cls._instances[key] = cls(platform, crawler_type)
            return cls._instances[key]

    @classmethod
    def flush_all(cls):
        """
        刷新所有Excel存储实例并保存到文件
        应在爬虫执行结束时调用
        """
        with cls._lock:
            for key, instance in cls._instances.items():
                try:
                    instance.flush()
                    utils.logger.info(f"[ExcelStoreBase] Flushed instance: {key}")
                except Exception as e:
                    utils.logger.error(f"[ExcelStoreBase] Error flushing {key}: {e}")
            cls._instances.clear()

    def __init__(self, platform: str, crawler_type: str = "search"):
        """
        初始化Excel存储

        Args:
            platform: 平台名称 (xhs, dy, ks等)
            crawler_type: 爬虫类型 (search, detail, creator)
        """
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "导出Excel需要openpyxl库。"
                "请使用: pip install openpyxl"
            )

        super().__init__()
        self.platform = platform
        self.crawler_type = crawler_type

        # 创建数据目录
        if config.SAVE_DATA_PATH:
            self.data_dir = Path(config.SAVE_DATA_PATH) / platform
        else:
            self.data_dir = Path("data") / platform
        self.data_dir.mkdir(parents=True, exist_ok=True)

        # 初始化工作簿
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # 移除默认工作表

        # 创建工作表
        self.contents_sheet = self.workbook.create_sheet("Contents")
        self.comments_sheet = self.workbook.create_sheet("Comments")
        self.creators_sheet = self.workbook.create_sheet("Creators")

        # 跟踪是否已写入表头
        self.contents_headers_written = False
        self.comments_headers_written = False
        self.creators_headers_written = False
        self.contacts_headers_written = False
        self.dynamics_headers_written = False

        # 可选工作表,适用于需要的平台(如Bilibili)
        self.contacts_sheet = None
        self.dynamics_sheet = None

        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = self.data_dir / f"{platform}_{crawler_type}_{timestamp}.xlsx"

        utils.logger.info(f"[ExcelStoreBase] Initialized Excel export to: {self.filename}")

    def _apply_header_style(self, sheet, row_num: int = 1):
        """
        应用表头行格式

        Args:
            sheet: 工作表对象
            row_num: 表头行号 (默认: 1)
        """
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in sheet[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

    def _auto_adjust_column_width(self, sheet):
        """
        根据内容自动调整列宽

        Args:
            sheet: 工作表对象
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass

            # 设置宽度,有最小/最大限制
            adjusted_width = min(max(max_length + 2, 10), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _write_headers(self, sheet, headers: List[str]):
        """
        向工作表写入表头

        Args:
            sheet: 工作表对象
            headers: 表头名称列表
        """
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        self._apply_header_style(sheet)

    def _write_row(self, sheet, data: Dict[str, Any], headers: List[str]):
        """
        向工作表写入数据行

        Args:
            sheet: 工作表对象
            data: 数据字典
            headers: 表头名称列表(定义列顺序)
        """
        row_num = sheet.max_row + 1

        for col_num, header in enumerate(headers, 1):
            value = data.get(header, "")

            # 处理不同的数据类型
            if isinstance(value, (list, dict)):
                value = str(value)
            elif value is None:
                value = ""

            cell = sheet.cell(row=row_num, column=col_num, value=value)

            # 应用基本格式
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    async def store_content(self, content_item: Dict):
        """
        将内容数据存储到Excel

        Args:
            content_item: 内容数据字典
        """
        # 定义表头(根据平台自定义)
        headers = list(content_item.keys())

        # 首次写入表头
        if not self.contents_headers_written:
            self._write_headers(self.contents_sheet, headers)
            self.contents_headers_written = True

        # 写入数据行
        self._write_row(self.contents_sheet, content_item, headers)

        # 从各种可能的字段名获取ID
        content_id = content_item.get('note_id') or content_item.get('aweme_id') or content_item.get('video_id') or content_item.get('content_id') or 'N/A'
        utils.logger.info(f"[ExcelStoreBase] 已存储内容到Excel: {content_id}")

    async def store_comment(self, comment_item: Dict):
        """
        将评论数据存储到Excel

        Args:
            comment_item: 评论数据字典
        """
        # Define headers
        headers = list(comment_item.keys())

        # Write headers if first time
        if not self.comments_headers_written:
            self._write_headers(self.comments_sheet, headers)
            self.comments_headers_written = True

        # Write data row
        self._write_row(self.comments_sheet, comment_item, headers)

        utils.logger.info(f"[ExcelStoreBase] Stored comment to Excel: {comment_item.get('comment_id', 'N/A')}")

    async def store_creator(self, creator: Dict):
        """
        将创作者数据存储到Excel

        Args:
            creator: 创作者数据字典
        """
        # 定义表头
        headers = list(creator.keys())

        # 首次写入表头
        if not self.creators_headers_written:
            self._write_headers(self.creators_sheet, headers)
            self.creators_headers_written = True

        # 写入数据行
        self._write_row(self.creators_sheet, creator, headers)

        utils.logger.info(f"[ExcelStoreBase] 已存储创作者到Excel: {creator.get('user_id', 'N/A')}")

    async def store_contact(self, contact_item: Dict):
        """
        将联系数据存储到Excel(适用于Bilibili等平台)

        Args:
            contact_item: 联系数据字典
        """
        # 如果不存在则创建联系人工作表
        if self.contacts_sheet is None:
            self.contacts_sheet = self.workbook.create_sheet("Contacts")

        # 定义表头
        headers = list(contact_item.keys())

        # 首次写入表头
        if not self.contacts_headers_written:
            self._write_headers(self.contacts_sheet, headers)
            self.contacts_headers_written = True

        # 写入数据行
        self._write_row(self.contacts_sheet, contact_item, headers)

        utils.logger.info(f"[ExcelStoreBase] 已存储联系人到Excel: up_id={contact_item.get('up_id', 'N/A')}, fan_id={contact_item.get('fan_id', 'N/A')}")

    async def store_dynamic(self, dynamic_item: Dict):
        """
        将动态数据存储到Excel(适用于Bilibili等平台)

        Args:
            dynamic_item: 动态数据字典
        """
        # 如果不存在则创建动态工作表
        if self.dynamics_sheet is None:
            self.dynamics_sheet = self.workbook.create_sheet("Dynamics")

        # 定义表头
        headers = list(dynamic_item.keys())

        # 首次写入表头
        if not self.dynamics_headers_written:
            self._write_headers(self.dynamics_sheet, headers)
            self.dynamics_headers_written = True

        # 写入数据行
        self._write_row(self.dynamics_sheet, dynamic_item, headers)

        utils.logger.info(f"[ExcelStoreBase] 已存储动态到Excel: {dynamic_item.get('dynamic_id', 'N/A')}")

    def flush(self):
        """
        保存工作簿到文件
        """
        try:
            # 为所有工作表自动调整列宽
            self._auto_adjust_column_width(self.contents_sheet)
            self._auto_adjust_column_width(self.comments_sheet)
            self._auto_adjust_column_width(self.creators_sheet)
            if self.contacts_sheet is not None:
                self._auto_adjust_column_width(self.contacts_sheet)
            if self.dynamics_sheet is not None:
                self._auto_adjust_column_width(self.dynamics_sheet)

            # 移除空工作表(仅表头行)
            if self.contents_sheet.max_row == 1:
                self.workbook.remove(self.contents_sheet)
            if self.comments_sheet.max_row == 1:
                self.workbook.remove(self.comments_sheet)
            if self.creators_sheet.max_row == 1:
                self.workbook.remove(self.creators_sheet)
            if self.contacts_sheet is not None and self.contacts_sheet.max_row == 1:
                self.workbook.remove(self.contacts_sheet)
            if self.dynamics_sheet is not None and self.dynamics_sheet.max_row == 1:
                self.workbook.remove(self.dynamics_sheet)

            # 检查是否还有工作表
            if len(self.workbook.sheetnames) == 0:
                utils.logger.info(f"[ExcelStoreBase] 没有数据要保存,跳过文件创建: {self.filename}")
                return

            # 保存工作簿
            self.workbook.save(self.filename)
            utils.logger.info(f"[ExcelStoreBase] Excel文件保存成功: {self.filename}")

        except Exception as e:
            utils.logger.error(f"[ExcelStoreBase] Error saving Excel file: {e}")
            raise
